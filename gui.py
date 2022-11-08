# -*- coding: utf-8 -*-

################################################################################
## Form generated from reading UI file 'guizeSQOx.ui'
##
## Created by: Qt User Interface Compiler version 5.14.1
##
## WARNING! All changes made in this file will be lost when recompiling UI file!
################################################################################

from cmath import inf
from unicodedata import name
from PySide2.QtCore import (QCoreApplication, QMetaObject, QObject, QPoint,
    QRect, QSize, QUrl, Qt)
from PySide2.QtGui import (QBrush, QColor, QConicalGradient, QCursor, QFont,
    QFontDatabase, QIcon, QLinearGradient, QPalette, QPainter, QPixmap,
    QRadialGradient)
from PySide2.QtWidgets import *
import openpyxl
import csv
import win32api, win32con, win32gui, ctypes
from regenerate import *

from ui_columnwidget import ColumnWidget
class Ui_MainWindow(object):
    filestr = ""
    minmaxes = []
    columns = []
    def setupUi(self, MainWindow):
        if MainWindow.objectName():
            MainWindow.setObjectName(u"xls regenerator")
        MainWindow.resize(800, 186)
        self.centralwidget = QWidget(MainWindow)
        self.centralwidget.setObjectName(u"centralwidget")
        self.scrollArea = QScrollArea(self.centralwidget)
        self.scrollArea.setObjectName(u"scrollArea")
        self.scrollArea.setGeometry(QRect(150, 10, 641, 166))
        self.scrollArea.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scrollArea.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.scrollArea.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.scrollArea.setLayout(QHBoxLayout())
        self.scrollArea.setMinimumSize(QSize(0, 0))
        self.scrollArea.setMaximumSize(QSize(16777215, 16777215))
        self.scrollArea.setWidgetResizable(True)
        self.scrollAreaWidgetContents = QWidget()
        self.scrollAreaWidgetContents.setObjectName(u"scrollAreaWidgetContents")
        self.scrollAreaWidgetContents.setGeometry(QRect(0, 0, 639, 112))
        self.column_frames = []#  = Column_widget(self.scrollArea)
        self.status = QLabel(self.centralwidget)
        self.status.setObjectName(u"status")
        self.status.setGeometry(QRect(15, 50, 131, 123))
        self.substatus = QLabel(self.centralwidget)
        self.substatus.setObjectName(u"status")
        self.substatus.setGeometry(QRect(15, 60, 131, 123))
        self.substatus.show()

        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.pushButton = QPushButton(self.centralwidget)
        self.pushButton.setObjectName(u"pushButton")
        self.pushButton.setGeometry(QRect(10, 10, 131, 23))
        self.pushButton_2 = QPushButton(self.centralwidget)
        self.pushButton_2.setObjectName(u"pushButton_2")
        self.pushButton_2.setGeometry(QRect(10, 40, 131, 23))
        MainWindow.setCentralWidget(self.centralwidget)
        self.retranslateUi(MainWindow)
        QMetaObject.connectSlotsByName(MainWindow)
    # setupUi

    def retranslateUi(self, MainWindow):
        #MainWindow.setWindowTitle(QCoreApplication.translate("MainWindow", u"MainWindow", None))
        self.pushButton.setText(QCoreApplication.translate("MainWindow", u"Wczytaj plik", None))
        self.pushButton_2.setText(QCoreApplication.translate("MainWindow", u"Zapisz plik", None))
        self.pushButton.clicked.connect(self.pushButton_clicked)
        self.pushButton_2.clicked.connect(self.pushButton_2_clicked)
    # retranslateUi

    def pushButton_clicked(self):
        for i in range(len(self.column_frames)):
            self.column_frames[i].deleteLater()
        self.filestr = QFileDialog.getOpenFileName(None, 'Wczytaj plik', '.\\', "Excel FIles (*.xls *.xlsx);;CSV files (*.csv)")[0]
        self.status.setText(QCoreApplication.translate("MainWindow", u"Wczytywanie pliku... ",None))
        QCoreApplication.processEvents()
        if self.filestr != "":
            if self.filestr[-3:] == "csv":
                with open(self.filestr, 'r') as csvfile:
                    self.data = list(csv.reader(csvfile, delimiter=';'))
                self.data = list(self.data)
                for i in range(len(self.data[0])):
                    self.add_column()
                    self.column_frames[-1].Column_name.setText(self.data[0][i])
                    self.substatus.setText(QCoreApplication.translate("MainWindow", self.data[0][i],None))
                    QCoreApplication.processEvents()
                    min, max = self.minmax(i)
                    self.column_frames[-1].MinVal.setText(str(min))
                    self.column_frames[-1].MaxVal.setText(str(max))
                    self.readcolumn(i)
            elif self.filestr[-3:] == "xls" or self.filestr[-4:] == "xlsx":
                self.data = openpyxl.load_workbook(self.filestr)
                self.sheet = self.data[self.data.sheetnames[0]]
                self.column_frames = []
                for i in range(1, self.sheet.max_column + 1):
                    self.add_column()
                    self.column_frames[-1].Column_name.setText(self.sheet.cell(row=1, column=i).value)
                    self.substatus.setText(QCoreApplication.translate("MainWindow", self.sheet.cell(row=1, column=i).value,None))
                    QCoreApplication.processEvents()
                    min, max = self.minmax(i)
                    self.column_frames[-1].MinVal.setText(str(min))
                    self.column_frames[-1].MaxVal.setText(str(max))
                    self.readcolumn(i)
        self.status.setText(QCoreApplication.translate("MainWindow", u"",None))
        self.substatus.setText(QCoreApplication.translate("MainWindow", u"",None))
        QCoreApplication.processEvents()

    def readcolumn(self, column):
        values = list()
        if self.filestr[-3] == "csv":
            for i,row in enumerate(self.data):
                if i == 0:
                    continue
                try:
                    values.append(float(row[column]))
                except:
                    pass
        elif self.filestr[-3:] == "xls" or self.filestr[-4:] == "xlsx":
            for row_cell in range(self.sheet.max_row):
                try:
                    values.append(float(self.sheet.cell(row=row_cell+1, column=column).value))
                except:
                    pass
        self.columns.append(values)

    def pushButton_2_clicked(self):
        if self.filestr != "":
            self.status.setText(QCoreApplication.translate("MainWindow", u"Regeneracja...",None))
            QCoreApplication.processEvents()
            if self.filestr[-3:] == "csv":
                with open(self.filestr, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile, delimiter=';')
                    for i in range(len(self.data[0])):
                        for j in range(len(self.data)):
                            try:
                                if (float(self.data[j][i]) > float(self.column_frames[i].MaxVal.text())
                                    or float(self.data[j][i]) < float(self.column_frames[i].MinVal.text())
                                    or column.check_box.isChecked()):
                                    self.data[j][i] = str(self.method(i,j))
                            except:
                                pass
                    self.status.setText(QCoreApplication.translate("MainWindow", u"Zapisywanie...",None))
                    QCoreApplication.processEvents()
                    writer.writerows(self.data)
            elif self.filestr[-3:] == "xls" or self.filestr[-4:] == "xlsx":
                for i,column in enumerate(self.column_frames):
                    for cell in range(self.sheet.max_row):
                        try:
                            a = float(self.sheet.cell(row=cell+1, column=i+1).value)
                            if (float(self.sheet.cell(row=cell+1, column=i+1).value) > float(column.MaxVal.text())
                                or float(self.sheet.cell(row=cell+1, column=i+1).value) < float(column.MinVal.text())
                                or column.checkBox.isChecked()):
                                self.sheet.cell(row=cell+1, column=i+1).value = self.method(i,cell)
                        except:
                            pass      
                self.status.setText(QCoreApplication.translate("MainWindow", u"Zapisywanie...",None))
                self.data.save(self.filestr)
        self.status.setText(QCoreApplication.translate("MainWindow", u"",None))
        self.substatus.setText(QCoreApplication.translate("MainWindow", u"",None))
        QCoreApplication.processEvents()

    def minmax(self, column):
        if column in range(len(self.minmaxes)): return self.minmaxes[column]
        values = []
        if self.filestr[-3:] == "csv":
            for i,row in enumerate(self.data):
                if i == 0:
                    continue
                try:
                    values.append(float(row[column]))
                except:
                    pass
        if self.filestr[-3:] == "xls" or self.filestr[-4:] == "xlsx":
            for cell in range(self.sheet.max_row):
                try:
                    values.append(float(self.sheet.cell(row=cell+1, column=column).value))
                except:
                    pass
        if values:
            self.minmaxes.append([min(values), max(values)])
            return min(values), max(values)
        self.minmaxes.append([-inf, inf])
        return -inf, inf

    def method(self, column, cell):
        try:
            min, max = float(self.column_frames[column].MinVal.text()), float(self.column_frames[column].MaxVal.text())
        except:
            min, max = -inf, inf
        try:
            maxchange = float(self.column_frames[column].MaxChangeVal.text())
        except:
            maxchange = 0
        try:
            maxchangerequired = float(self.column_frames[column].checkBox.isChecked())
        except:
            maxchangerequired = False
        values = self.columns[column]
        result = 0
        if self.column_frames[column].Method.currentText() == "Średnia":
            self.substatus.setText(QCoreApplication.translate("MainWindow", "Średnia " + self.column_frames[column].Column_name.text(),None))
            QCoreApplication.processEvents()
            result = calculate_medium([min,max],values,maxchange,maxchangerequired)
        elif self.column_frames[column].Method.currentText() == "Poprzedzająca":
            self.substatus.setText(QCoreApplication.translate("MainWindow", "Poprzedzająca " + self.column_frames[column].Column_name.text(),None))
            QCoreApplication.processEvents()
            result = find_good_previous_value([min,max],values,cell,maxchange,maxchangerequired)
        elif self.column_frames[column].Method.currentText() == "Następująca":
            self.substatus.setText(QCoreApplication.translate("MainWindow", "Następująca " + self.column_frames[column].Column_name.text(),None))
            QCoreApplication.processEvents()
            result = find_good_next_value([min,max],values,cell,maxchange,maxchangerequired)
        elif self.column_frames[column].Method.currentText() == "Pośrednia":
            self.substatus.setText(QCoreApplication.translate("MainWindow", "Pośrednia " + self.column_frames[column].Column_name.text(),None))
            QCoreApplication.processEvents()
            prev, next = find_good_previous_value([min,max],values,cell,maxchange,maxchangerequired), find_good_next_value([min,max],values,cell,maxchange,maxchangerequired)
            result = (prev + next) / 2
        if result > max:
            result = max
        elif result < min:
            result = min
        self.columns[column][cell] = result
        return result

    def add_column(self):
        coords = [0,0]
        if (self.column_frames):
            last_frame = self.column_frames[-1]
            coords[0] = last_frame.x() + last_frame.width()
        self.column_frames.append(ColumnWidget(self.scrollAreaWidgetContents))
        self.column_frames[-1].setGeometry(QRect(coords[0], coords[1], 181, 146))
        self.column_frames[-1].show()
        oldwidth = self.scrollAreaWidgetContents.width()
        if oldwidth < self.column_frames[-1].x() * len(self.column_frames):
            width = self.column_frames[-1].width() * len(self.column_frames)
            heigth = self.scrollAreaWidgetContents.height()
            self.scrollAreaWidgetContents.setGeometry(QRect(0, 0, width, heigth))
            self.scrollAreaWidgetContents.setMinimumSize(QSize(width, heigth))
            self.scrollAreaWidgetContents.setMaximumSize(QSize(width, heigth))
        self.scrollAreaWidgetContents.update()
        

        


if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    MainWindow = QMainWindow()
    ui = Ui_MainWindow()
    MainWindow.setWindowTitle("xls regenerator")
    try:
        handle = win32api.GetModuleHandle(None)
        icon = win32api.LoadResource(handle, win32con.RT_ICON, 1)
        pixmap = QPixmap()
        pixmap.loadFromData(icon)
        icon = QIcon(pixmap)  
        MainWindow.setWindowIcon(icon)
    except:
        pass
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())